from typing import Dict, Any, List
import io
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelService:
    # KPMG Blue color scheme
    KPMG_BLUE = "00338D"
    KPMG_LIGHT_BLUE = "0091DA"
    HEADER_FILL = "00338D"
    
    def generate_excel_report(self, df: pd.DataFrame, stats: Dict[str, Any], insights: List[str]) -> io.BytesIO:
        """Generate comprehensive Excel report with multiple sheets"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, df, insights)
        self._create_data_quality_sheet(wb, df)
        self._create_statistics_sheet(wb, df, stats)
        self._create_correlation_sheet(wb, df)
        self._create_raw_data_sheet(wb, df)
        
        # Save to BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _create_summary_sheet(self, wb: Workbook, df: pd.DataFrame, insights: List[str]):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Data Analysis Report - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Dataset Overview
        row = 3
        ws[f'A{row}'] = "Dataset Overview"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.KPMG_BLUE)
        row += 1
        
        overview_data = [
            ["Total Records", len(df)],
            ["Total Columns", len(df.columns)],
            ["Total Missing Values", df.isnull().sum().sum()],
            ["Missing Percentage", f"{(df.isnull().sum().sum() / df.size * 100):.2f}%"]
        ]
        
        for metric, value in overview_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Key Insights
        row += 1
        ws[f'A{row}'] = "Key Insights"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.KPMG_BLUE)
        row += 1
        
        for i, insight in enumerate(insights[:5], 1):
            ws[f'A{row}'] = f"{i}. {insight[:100]}..."
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_data_quality_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create data quality report sheet"""
        ws = wb.create_sheet("Data Quality")
        
        # Title
        ws['A1'] = "Data Quality Report"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ["Column", "Data Type", "Missing Count", "Missing %", "Unique Values"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.KPMG_LIGHT_BLUE, end_color=self.KPMG_LIGHT_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        row = 4
        for col in df.columns:
            missing_count = df[col].isnull().sum()
            missing_pct = (missing_count / len(df) * 100)
            
            ws.cell(row=row, column=1, value=str(col))
            ws.cell(row=row, column=2, value=str(df[col].dtype))
            ws.cell(row=row, column=3, value=missing_count)
            ws.cell(row=row, column=4, value=f"{missing_pct:.2f}%")
            ws.cell(row=row, column=5, value=df[col].nunique())
            
            # Highlight high missing values
            if missing_pct > 20:
                for col_idx in range(1, 6):
                    ws.cell(row=row, column=col_idx).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
    
    def _create_statistics_sheet(self, wb: Workbook, df: pd.DataFrame, stats: Dict[str, Any]):
        """Create statistical summary sheet"""
        ws = wb.create_sheet("Statistical Summary")
        
        # Title
        ws['A1'] = "Statistical Summary - Numeric Variables"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        ws.merge_cells('A1:F1')
        
        # Get numeric columns
        numeric_df = df.select_dtypes(include=[np.number])
        
        if not numeric_df.empty:
            # Headers
            headers = ["Column", "Mean", "Median", "Std Dev", "Min", "Max"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=self.KPMG_LIGHT_BLUE, end_color=self.KPMG_LIGHT_BLUE, fill_type="solid")
            
            # Data
            row = 4
            for col in numeric_df.columns:
                ws.cell(row=row, column=1, value=str(col))
                ws.cell(row=row, column=2, value=round(numeric_df[col].mean(), 2))
                ws.cell(row=row, column=3, value=round(numeric_df[col].median(), 2))
                ws.cell(row=row, column=4, value=round(numeric_df[col].std(), 2))
                ws.cell(row=row, column=5, value=round(numeric_df[col].min(), 2))
                ws.cell(row=row, column=6, value=round(numeric_df[col].max(), 2))
                row += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _create_correlation_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create correlation matrix sheet"""
        ws = wb.create_sheet("Correlation Matrix")
        
        # Title
        ws['A1'] = "Correlation Matrix - Numeric Variables"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        
        numeric_df = df.select_dtypes(include=[np.number])
        
        if not numeric_df.empty and numeric_df.shape[1] >= 2:
            corr_matrix = numeric_df.corr()
            
            # Write correlation matrix
            for r_idx, row in enumerate(dataframe_to_rows(corr_matrix, index=True, header=True), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Format correlation values
                    if isinstance(value, (int, float)) and r_idx > 4 and c_idx > 1:
                        cell.number_format = '0.00'
                        # Color code correlations
                        if abs(value) > 0.7:
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif abs(value) > 0.4:
                            cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            
            # Format headers
            for col_idx in range(1, len(corr_matrix.columns) + 2):
                ws.cell(row=3, column=col_idx).font = Font(bold=True)
                ws.cell(row=4, column=col_idx).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=4, column=col_idx).fill = PatternFill(start_color=self.KPMG_LIGHT_BLUE, end_color=self.KPMG_LIGHT_BLUE, fill_type="solid")
    
    def _create_raw_data_sheet(self, wb: Workbook, df: pd.DataFrame):
        """Create raw data sheet (limited to first 1000 rows)"""
        ws = wb.create_sheet("Raw Data")
        
        # Title
        ws['A1'] = f"Raw Data (First {min(1000, len(df))} rows)"
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_FILL, end_color=self.HEADER_FILL, fill_type="solid")
        
        # Write data (limit to 1000 rows for performance)
        limited_df = df.head(1000)
        
        for r_idx, row in enumerate(dataframe_to_rows(limited_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format headers
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.KPMG_LIGHT_BLUE, end_color=self.KPMG_LIGHT_BLUE, fill_type="solid")
